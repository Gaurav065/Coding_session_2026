# Databricks notebook source
# MAGIC %md
# MAGIC # Row Count Verification Framework
# MAGIC
# MAGIC **Purpose:** For any set of loads, compare row counts of base & target tables across
# MAGIC the three environments (`migration_prd`, `cons_prd`, `td_edwpc`) and produce an Excel report.
# MAGIC
# MAGIC **How to use:**
# MAGIC 1. Edit the `LOADS_CONFIG` dict in Cell 2 — add/remove loads and their base/target tables.
# MAGIC 2. Run the notebook top to bottom.
# MAGIC 3. The Excel report appears at the path printed in the last cell — download it from there.
# MAGIC
# MAGIC **Behaviour:**
# MAGIC - One table failure does NOT stop the run. The error is captured and the loop continues.
# MAGIC - The final report has three sheets: `Summary` (per-load verdict), `Detail` (every table), `Issues` (mismatches + errors only).
 
# COMMAND ----------
 
# MAGIC %md
# MAGIC ## Cell 1 — Imports & environment configuration
# MAGIC Edit `ENVIRONMENTS` and `CATALOG_SWAP` only if your environment names change.
 
# COMMAND ----------
 
import pandas as pd
from datetime import datetime
from concurrent.futures import ThreadPoolExecutor, as_completed
import traceback
 
# ----- The three environments we want to compare -----
# Logical name -> the catalog token used in the FQN.
ENVIRONMENTS = {
    "MIGRATION_PRD": "1dp_migration_prd",
    "CONS_PRD":      "1dp_cons_prd",
    "TD_EDWPC":      "1dp_migration_td_edwpc",
}
 
# ----- How catalog tokens map across environments -----
# Key   = the catalog token as written in the LOADS_CONFIG below.
# Value = what to substitute for each environment.
# Add new top-level catalogs here if your tables use them.
CATALOG_SWAP = {
    "1dp_migration_prd": {
        "MIGRATION_PRD": "1dp_migration_prd",
        "CONS_PRD":      "1dp_cons_prd",
        "TD_EDWPC":      "1dp_migration_td_edwpc",
    },
    # NOTE: these src catalog names are best-guesses — confirm with the team
    # what the parallel src catalogs are in cons and edwpc, then update here.
    "1dp_src_prd": {
        "MIGRATION_PRD": "1dp_src_prd",
        "CONS_PRD":      "1dp_cons_src_prd",
        "TD_EDWPC":      "1dp_src_td_edwpc",
    },
}
 
# ----- Run options -----
MAX_PARALLEL_COUNTS = 8     # How many COUNT(*) queries to run in parallel
OUTPUT_DIR          = "/dbfs/FileStore/row_count_reports"
 
# COMMAND ----------
 
# MAGIC %md
# MAGIC ## Cell 2 — Define the loads to verify
# MAGIC This is the only cell you normally need to edit. Add a load with its base tables and target tables.
# MAGIC Use the **migration_prd** fully-qualified name for each table — the framework will derive the names for the other two environments automatically.
 
# COMMAND ----------
 
LOADS_CONFIG = {
 
    "SALES_ORDER_INTEGRATION_SOBM": {
        "base": [
            "1dp_src_prd.tas_core.core_sales_order_business",
            "1dp_src_prd.tas_core.core_sales_order_header",
            "1dp_src_prd.tas_core.core_sales_order_item",
            "1dp_src_prd.tas_core.core_sales_order_status",
            "1dp_src_prd.tas_core.core_sales_order_sched_line",
            "1dp_src_prd.tas_core.core_rlip_clip_reporting",
            "1dp_src_prd.tas_core.core_sales_order_partner",
            "1dp_migration_prd.reference_base_t.material_nxp",
            "1dp_migration_prd.reference_base_t.management_organization_hierarchy_tv",
            "1dp_migration_prd.reference_base_t.management_organization_unit_tv",
            "1dp_migration_prd.reference_base_t.crm_cacc_nxp",
            "1dp_src_prd.tas_core.core_base_region",
            "1dp_src_prd.tas_core.core_consolidated_account",
            "1dp_src_prd.tas_core.core_customer",
            "1dp_src_prd.tas_core.core_customer_characts",
            "1dp_src_prd.tas_core.core_customer_delivery_prio",
            "1dp_src_prd.tas_core.core_industry_segment",
            "1dp_src_prd.tas_core.core_industry_sub_segment",
            "1dp_src_prd.tas_core.core_pd_key_account",
            "1dp_migration_prd.eim_base_reference_t.location_nxp_country_region",
            "1dp_src_prd.tas_core.core_customer_sales",
        ],
        "target": [
            "1dp_migration_prd.SC_PLANNING_BASE_T.SALES_ORDER_INTEGRATION_SOFT_LAUNCH",
        ],
    },
 
    "SALES_ORDER_INTEGRATION": {
        "base": [
            "1dp_src_prd.tas_core.core_sales_order_business",
            "1dp_src_prd.tas_core.core_sales_order_header",
            "1dp_src_prd.tas_core.core_sales_order_item",
            "1dp_src_prd.tas_core.core_sales_order_status",
            "1dp_src_prd.tas_core.core_sales_order_sched_line",
            "1dp_src_prd.tas_core.core_rlip_clip_reporting",
            "1dp_src_prd.tas_core.core_sales_order_partner",
            "1dp_migration_prd.edw_security.row_group_access",
            "1dp_migration_prd.edw_security.row_sec_group",
            "1dp_migration_prd.eim_base_reference_t.location_nxp_country_region",
            "1dp_migration_prd.reference_base_t.crm_cacc_nxp",
            "1dp_migration_prd.reference_base_t.management_organization_hierarchy_tv",
            "1dp_migration_prd.reference_base_t.management_organization_unit_tv",
            "1dp_migration_prd.reference_base_t.material_nxp",
            "1dp_migration_prd.sc_planning_base_t.sales_order_integration",
            "1dp_src_prd.tas_core.core_base_region",
            "1dp_src_prd.tas_core.core_consolidated_account",
            "1dp_src_prd.tas_core.core_customer",
            "1dp_src_prd.tas_core.core_customer_characts",
            "1dp_src_prd.tas_core.core_customer_delivery_prio",
            "1dp_src_prd.tas_core.core_customer_sales",
            "1dp_src_prd.tas_core.core_industry_segment",
            "1dp_src_prd.tas_core.core_industry_sub_segment",
            "1dp_src_prd.tas_core.core_pd_key_account",
        ],
        "target": [
            "1dp_migration_prd.SC_PLANNING_BASE_T.SALES_ORDER_INTEGRATION",
        ],
    },
 
    "IWH_WAREHOUSE_INVENTORY_SUMMARY": {
        "base": [
            "1dp_migration_prd.MARCCS.V_IM29_EBI_REPORT_ATBK",
            "1dp_migration_prd.MARCCS.V_IM29_EBI_REPORT_ATKH",
            "1dp_migration_prd.MARCCS.V_IM29_EBI_REPORT_ATKL",
            "1dp_migration_prd.MARCCS.V_IM29_EBI_REPORT_ATKLE",
            "1dp_migration_prd.MARCCS.V_IM29_EBI_REPORT_ATTJ",
            "1dp_migration_prd.MARCCS.V_IM29_EBI_REPORT_ATTX",
            "1dp_migration_prd.MARCCS.V_IM29_EBI_REPORT_ATTXE",
            "1dp_migration_prd.MARCCS.V_IM29_EBI_REPORT_ICN8",
        ],
        "target": [
            "1dp_migration_prd.SC_PLANNING_BASE_T.IWH_WAREHOUSE_INVENTORY_SUMMARY",
        ],
    },
 
    "IWH_WAREHOUSE_INVENTORY": {
        "base": [
            "1dp_migration_prd.reference_base_t.calendar_reference",
        ],
        "target": [
            "1dp_migration_prd.SC_PLANNING_BASE_T.IWH_WAREHOUSE_INVENTORY",
        ],
    },
 
}
 
# COMMAND ----------
 
# MAGIC %md
# MAGIC ## Cell 3 — Core framework (no edits needed)
 
# COMMAND ----------
 
def resolve_env_fqn(table_fqn: str, env_key: str) -> str:
    """Swap the first catalog token of table_fqn for the equivalent in env_key."""
    head, sep, rest = table_fqn.partition(".")
    if not sep:
        return table_fqn
    if head not in CATALOG_SWAP:
        # Unknown catalog token — leave the FQN unchanged so the row at least runs
        # (it will just compare the same physical table three times)
        return table_fqn
    return f"{CATALOG_SWAP[head][env_key]}.{rest}"
 
 
def safe_count(fqn: str):
    """Run COUNT(*) on a single table. Returns (count, error_message_or_None)."""
    try:
        result = spark.sql(f"SELECT COUNT(*) AS cnt FROM {fqn}").collect()
        return int(result[0]["cnt"]), None
    except Exception as e:
        # Trim the message — Spark errors can be huge
        msg = str(e).split("\n")[0]
        return None, msg[:300]
 
 
def build_work_items(loads_config):
    """Flatten the config into a list of (load, table_type, original_fqn, env_key, env_fqn) tasks."""
    items = []
    for load_name, parts in loads_config.items():
        for table_type, tables in parts.items():
            for table in tables:
                for env_key in ENVIRONMENTS.keys():
                    items.append({
                        "load":       load_name,
                        "type":       table_type.upper(),
                        "original":   table,
                        "env":        env_key,
                        "env_fqn":    resolve_env_fqn(table, env_key),
                    })
    return items
 
 
def run_verification(loads_config, max_workers=MAX_PARALLEL_COUNTS):
    """Run all COUNT queries in parallel, then collapse to one row per (load, type, table)."""
    items = build_work_items(loads_config)
    print(f"Starting verification: {len(items)} count queries across "
          f"{sum(len(t) for parts in loads_config.values() for t in parts.values())} tables…")
 
    # ----- Run all counts in parallel, capturing failures per item -----
    raw_results = {}  # keyed by (load, type, original, env)
    with ThreadPoolExecutor(max_workers=max_workers) as pool:
        future_to_item = {pool.submit(safe_count, it["env_fqn"]): it for it in items}
        completed = 0
        for fut in as_completed(future_to_item):
            it = future_to_item[fut]
            cnt, err = fut.result()
            raw_results[(it["load"], it["type"], it["original"], it["env"])] = (cnt, err, it["env_fqn"])
            completed += 1
            if completed % 10 == 0 or completed == len(items):
                print(f"  …{completed}/{len(items)} queries done")
 
    # ----- Collapse into one row per (load, type, original_table) -----
    seen = set()
    rows = []
    for it in items:
        key = (it["load"], it["type"], it["original"])
        if key in seen:
            continue
        seen.add(key)
 
        row = {
            "load_name":  it["load"],
            "table_type": it["type"],
            "table_name": it["original"],
        }
        errors = []
        counts = []
        for env_key in ENVIRONMENTS.keys():
            cnt, err, env_fqn = raw_results[(it["load"], it["type"], it["original"], env_key)]
            row[f"{env_key}_count"] = cnt
            row[f"{env_key}_fqn"]   = env_fqn
            counts.append(cnt)
            if err:
                errors.append(f"{env_key}: {err}")
 
        # ----- Verdict -----
        if any(c is None for c in counts):
            row["match_status"] = "ERROR"
            row["max_diff"]     = None
        elif len(set(counts)) == 1:
            row["match_status"] = "MATCH"
            row["max_diff"]     = 0
        else:
            row["match_status"] = "MISMATCH"
            row["max_diff"]     = max(counts) - min(counts)
        row["error_detail"] = " | ".join(errors) if errors else ""
        rows.append(row)
 
    df = pd.DataFrame(rows)
    # Stable column order
    ordered_cols = (
        ["load_name", "table_type", "table_name", "match_status", "max_diff"]
        + [f"{e}_count" for e in ENVIRONMENTS]
        + [f"{e}_fqn"   for e in ENVIRONMENTS]
        + ["error_detail"]
    )
    return df[ordered_cols]
 
 
def build_summary(detail_df):
    """One row per load, with per-status counts and a verdict."""
    g = detail_df.groupby("load_name")
    summary = g.agg(
        total_tables=("table_name",  "count"),
        matches=("match_status",     lambda s: (s == "MATCH").sum()),
        mismatches=("match_status",  lambda s: (s == "MISMATCH").sum()),
        errors=("match_status",      lambda s: (s == "ERROR").sum()),
    ).reset_index()
 
    def verdict(r):
        if r["errors"] > 0:                 return "HAS ERRORS"
        if r["mismatches"] > 0:             return "NEEDS REVIEW"
        return "ALL MATCH"
    summary["load_verdict"] = summary.apply(verdict, axis=1)
    return summary
 
 
def write_excel_report(detail_df, summary_df, output_path):
    """Write a multi-sheet Excel with conditional colouring for the status column."""
    import os
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
 
    issues_df = detail_df[detail_df["match_status"] != "MATCH"].reset_index(drop=True)
 
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        summary_df.to_excel(writer, sheet_name="Summary", index=False)
        detail_df.to_excel(writer,  sheet_name="Detail",  index=False)
        issues_df.to_excel(writer,  sheet_name="Issues",  index=False)
 
        # ----- Light formatting: colour the status column on each sheet -----
        from openpyxl.styles import PatternFill, Font, Alignment
 
        green  = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        red    = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        yellow = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        header_fill = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
 
        status_colors = {
            "MATCH": green, "MISMATCH": red, "ERROR": yellow,
            "ALL MATCH": green, "NEEDS REVIEW": red, "HAS ERRORS": yellow,
        }
 
        for sheet_name, df in [("Summary", summary_df), ("Detail", detail_df), ("Issues", issues_df)]:
            ws = writer.sheets[sheet_name]
 
            # Header row formatting
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
 
            # Find status columns (match_status / load_verdict) and colour cells
            status_col_idx = None
            for i, col_name in enumerate(df.columns, start=1):
                if col_name in ("match_status", "load_verdict"):
                    status_col_idx = i
                    break
            if status_col_idx is not None:
                for row_i in range(2, ws.max_row + 1):
                    cell = ws.cell(row=row_i, column=status_col_idx)
                    fill = status_colors.get(str(cell.value))
                    if fill:
                        cell.fill = fill
 
            # Autosize columns (approximate)
            for col_cells in ws.columns:
                max_len = 10
                col_letter = col_cells[0].column_letter
                for c in col_cells:
                    if c.value is not None:
                        max_len = max(max_len, min(60, len(str(c.value)) + 2))
                ws.column_dimensions[col_letter].width = max_len
 
            ws.freeze_panes = "A2"
 
    return output_path
 
# COMMAND ----------
 
# MAGIC %md
# MAGIC ## Cell 4 — Run the verification
 
# COMMAND ----------
 
detail_df  = run_verification(LOADS_CONFIG)
summary_df = build_summary(detail_df)
 
print("\n========== SUMMARY ==========")
print(summary_df.to_string(index=False))
 
print("\n========== ISSUES (mismatches + errors) ==========")
issues = detail_df[detail_df["match_status"] != "MATCH"]
if issues.empty:
    print("None — everything matched.")
else:
    print(issues[["load_name", "table_type", "table_name", "match_status",
                  "MIGRATION_PRD_count", "CONS_PRD_count", "TD_EDWPC_count",
                  "max_diff", "error_detail"]].to_string(index=False))
 
# COMMAND ----------
 
# MAGIC %md
# MAGIC ## Cell 5 — Pretty notebook view + Excel report
 
# COMMAND ----------
 
# In-notebook interactive view
display(spark.createDataFrame(detail_df))
 
# COMMAND ----------
 
# Write the Excel report
timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
report_path = f"{OUTPUT_DIR}/row_count_report_{timestamp}.xlsx"
write_excel_report(detail_df, summary_df, report_path)
 
# Build a Databricks /files/ URL so you can click to download
download_url = report_path.replace("/dbfs/FileStore/", "/files/")
print(f"\n✅ Report written: {report_path}")
print(f"   Download via:    {download_url}")
print(f"   Full URL:        https://<your-workspace>.cloud.databricks.com{download_url}")